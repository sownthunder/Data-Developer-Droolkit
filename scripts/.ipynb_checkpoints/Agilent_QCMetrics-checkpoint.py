# -*- coding: utf-8 -*-
"""
Created on Mon Feb 24 16:16:13 2020

@author: derbates

TABLES INPUTEED:
    - tblProdflow
    - PRODUCTS

==========================================
DETERMINES NUMBER OF "PfBatchID" over 
last 30 days, in 3 level bins 
(L1, L2, L3)
==========================================

"""

# IMPORT THE GOODS
import os, sys, time
from time import sleep
from pathlib import Path
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import numpy as np
from pandas import Series, DataFrame
import pyodbc
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
from ttkthemes import ThemedStyle
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox, filedialog, commondialog


class AgilentQCMetrics(): # {
    
    user_name = str(os.getlogin()) # get username 
    outbound_dir = "C:/data/outbound/"
    desktop_dir = "OneDrive - Agilent Technologies/Desktop" #C:/Users/derbates
    
    def __init__(self, root, the_logger): # {
        self.root = root
        self.root.title("QC Metrics")
        self.root.geometry('315x200+300+300')
        self.root.resizable(width=False, height=False)
        # Get/Set USERNAME & DESKTOP DIRECTRIES
        self.user_name_dir = os.path.join("C:/Users/", self.user_name)
        self.desktop_dir = os.path.join(self.user_name_dir, self.desktop_dir)
        print(self.user_name_dir)
        print(self.desktop_dir)
        self.create_gui(the_root = self.root)
    # }
    
    def create_gui(self, the_root): # {
        # TRY THE FOLLOWING
        try: # { 
            self.create_ttk_styles(the_root=the_root)
            self.create_label_frame(the_root=the_root)
            self.create_main_frame(the_root=the_root)
        # }
        except: # {
            pass
        # }
    # }
    
    def create_ttk_styles(self, the_root): # {
        # TRY THE FOLLOWING
        try: # {
            # CONFIGURE THE TTK STYLE
            self.style = ThemedStyle(the_root)
            # STYLE THEME
            self.style.set_theme("radiance")
        # }
        except:# { 
            pass
        # }
    # }
    
    def create_label_frame(self, the_root): # {
        self.labelframe = ttk.Frame(the_root)
        self.labelframe.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    # }
    
    def create_main_frame(self, the_root): # {
        # TRY THE FOLLOWING
        try: # {
            self.mainframe = ttk.Frame(self.labelframe)
            self.mainframe.pack(anchor=tk.CENTER, fill=tk.BOTH, expand=True)
            # END DATE LABEL
            ttk.Label(master=self.mainframe, text="Enter Date: \n(YYYY-MM-DD)"
                      ).place(x=20, y=20) #.grid(row=0, col=0, padx=10, pady=10) 
            # CREATE END DATE TK VAR TO HOLD STR
            self.end_date = tk.StringVar(master=the_root)
            # ENTRY FOR END DATE
            ttk.Entry(master=self.mainframe, textvariable=self.end_date
                      ).place(x=140, y=20) #.grid(row=0, col=1, padx=10, pady=10)
            # NUM OF DAYS LABEL
            ttk.Label(master=self.mainframe, text="# of days back: "
                      ).place(x=20, y=60)# .grid(row=1, col=0, padx=10, pady=10) 
            self.num_of_days = tk.IntVar(master=the_root, value=1)
            # NUM OF DAYS SPINBOX
            ttk.Spinbox(master=self.mainframe, to=100, from_=0, textvariable=self.num_of_days
                        ).place(x=140, y=55) #.grid(row=1, col=1, padx=10, pady=10)
            # CREATE STRING TO HOLD POSSIBLE FILENAME
            # [2020-03-08]\\self.period_str = tk.StringVar(master=the_root, value="Day")
            self.filename_str = tk.StringVar(master=the_root, value=str(self.end_date.get()))
            # FILENAME LABELS AND ENTRY BOX
            ttk.Label(master=self.mainframe, text="Filename:\n\tQCMetrics_"
                      ).place(x=20, y= 100)
            ttk.Entry(master=self.mainframe, textvariable=self.end_date, state=tk.DISABLED
                      ).place(x=140, y= 100)
            ttk.Label(master=self.mainframe, text=".xlsx"
                      ).place(x=275, y= 100)
            # BROWSE FOR EXPORT LOCATION BUTTON
            self.export_button = ttk.Button(master=self.mainframe, text="EXPORT TO DESKTOP",
                       command=self.determine_range, width=25).place(x=20, y = 140)
            # [2020-03-06]\\self.export_button.bind("<Return>", self.test_event)
            """
            # EXPORT LOCATION TK VAR TO HOLD STR
            self.export_directory = tk.StringVar(master=the_root, value="...")
            # EXPORT LOCATION ENTRY BOX
            ttk.Entry(master=self.mainframe, textvariable=self.export_directory,
                      width=40).place(x=20, y=145)
            """
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n" + str(errorMessage))
            print("\n" + typeE +
                             "\n" + fileE +
                             "\n" + lineE +
                             "\n" + messageE)
            messagebox.showerror(title="ERROR!",
                                 message=typeE +
                                         "\n" + fileE +
                                         "\n" + lineE +
                                         "\n" + messageE)
        # }
        else: # {
            pass
        # }
    # }
    
    def test_event(self, event): # {
        print(str(event))
    # }
    
    def determine_range(self): # {
        print("Determining...\n")
        #print(str(self.start_date.get()))
        print(str(self.end_date.get()))
        print(str(self.num_of_days.get()))
        #print(str(self.period_str.get()))
        # CALL RUN
        self.run(date_input=str(self.end_date.get()), day_range=int(self.num_of_days.get()))
    # }
    
    def run(self, date_input, day_range): # {
        # TRY THE FOLLOWING
        try: # {
            # CREATE DATE OFF INPUT
            the_date = pd.Timestamp(ts_input=str(date_input))
            # create day range
            # [2020-03-08]\\x_days_ago = the_date - pd.Timedelta(unit='D', value=int(day_range))
            x_days_ago = the_date - timedelta(days = int(the_date))
            df_x_range = pd.date_range(start=x_days_ago, end=the_date, freq='D')
            print(list(df_x_range))
            ###########################################################
            # CREATE METRICS TABLE FROM CLASS METHOD
            self.df_metrics_table = self.create_metrics_table()
            ###########################3###############################
            df_last_range = self.df_QCMetrics[str(x_days_ago):str(the_date)]
            print(len(df_last_range))
            value_list = ['1.0', '2.0', '3.0']
            # Grab DataFrame rows where column has certain values
            df_last_range = df_last_range[df_last_range['ProductLevel'].isin(value_list)]
            """
            GROUPBY MULTIPLE COLUMNS
            """
            df_daily_levels = pd.DataFrame(data=df_last_range.groupby([df_last_range.index, 'ProductLevel'])['PfBatchID'].count())
            # RENAME COLUMNS
            df_daily_levels.rename(columns={'PfBatchID': 'Count'}, inplace=True)
            print(df_daily_levels.info())
            # SORT INDEX
            df_daily_levels.sort_index(inplace=True)
            # UNSTACKED INDEX (remove layer?)
            df_daily_unstacked_1 = df_daily_levels.unstack(axis=-1)
            # LIST NUMBER OF COLUMN LEVELS
            print(len(df_daily_unstacked_1.columns.levels))
            ##############
            # DROP LEVEL #
            ##############
            df_daily_unstacked_1.columns = df_daily_unstacked_1.columns.droplevel()
            # FILL NA
            df_daily_unstacked_1.fillna(value=0, inplace=True)
            ############################################################################
            # WORK BOOK FUNCTIONS >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
            ############################################################################
            print(self.desktop_dir)
            # CREATE STR VAR for filename convention
            ts_str = str(the_date)[:10]
            # CREATE FILENAME VAR
            filename_var = str("QCMetrics_" + str(ts_str) + ".xlsx")
            # CREATE FULL WORKBOOK PATH
            workbook_path = os.path.join(self.desktop_dir, str(filename_var))
            print("WORKBOOK_PATH == " + str(workbook_path))
            # CREATE NEW WORKBOOK
            wb = Workbook()
            wb.save(workbook_path)
            # ADD SHEETS TO WORKBOOK
            # DESGINATE SHEET NAME AND POSITION
            sheet1 = wb.create_sheet('Table', 0)
            sheet2 = wb.create_sheet('Graphs', 1)
            # ACTIVATE WORKSHEET TO WRITE DATAFRAME
            active = wb['Table']
            # WRITE DATAFRAME TO ACTIVE WORKSHEET
            for x in dataframe_to_rows(df_daily_unstacked_1): # {
                active.append(x)
            # }
            # SAVE
            wb.save(filename_var)
            # CREATE LINE PLOT VARIABLE
            plot = df_daily_unstacked_1.plot()
            # CREATE AREA PLOT VARIABLE
            area_plot = df_daily_unstacked_1.plot(kind='area')
            # MATPLOTLIB figure for "line_plot"
            line_fig = plot.get_figure()
            # MATPLOTLIB figure for "area_plot"
            area_fig = area_plot.get_figure()
            ###################
            # CREATE TEMP DIR #
            ###################
            with tempfile.TemporaryDirectory() as directory_name: # {
                the_dir = Path(directory_name)
                print("TEMPORARY DIR == " + str(the_dir))
                line_img_path = os.path.join(the_dir, str(ts_str) + "_line_plot.png")
                area_img_path = os.path.join(the_dir, str(ts_str) + "_area_plot.png")
                
                # SAVE LINE PLOT
                line_fig.savefig(line_img_path)
                # SAVE AREA PLOT
                area_fig.savefig(area_img_path)
                # ACTIVATE WORKSHEET
                active = wb['Graphs']
                # Insert Plot into Worksheet
                # Select active sheet and cell reference
                img_line = Image(line_img_path)
                active.add_image(img_line, 'A1')
                # Insert Plot into worksheet
                # Select active sheet and cell reference
                img_area= Image(area_img_path)
                active.add_image(img_area, 'H1')
                # SAVE WORKBOOK
                wb.save(filename_str)
            # }
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
    # }
    
    # [2020-03-08]\\
    """
    def run(self, day_range): # {
        # TRY THE FOLLOWING
        try: # {
            # [2020-02-28]\\self.time_unit = time_unit
            # [2020-02-28]\\self.time_value = time_value
            self.day_range = day_range
            # get/set current date variable
            # [2020-03-06]\\the_date = pd.Timestamp.now()
            the_date = pd.Timestamp(ts_input=str(self.end_date.get()))
            # create variable for a month ago
            one_month_ago = the_date - timedelta(days = int(day_range))
            # [2020-02-28]\\one_month_ago = the_date - pd.Timedelta(unit=str(self.time_unit), value=str(self.time_value))
            print(type(one_month_ago))
            # CREATE METRICS TABLE FROM CLASS METHOD
            self.df_metrics_table = self.create_metrics_table()
            #############################
            # create .csv with no drops #
            #############################
            self.df_metrics_table.to_csv(os.path.join(self.outbound_dir, "df_metrics_noDrop.csv"), index=True)
            # DROP ROWS WITH PRODUCT LEVEL && QCDATE = NONE
            self.df_metrics_table.dropna(axis=0, subset=['QCDate', 'ProductLevel'], how='any', inplace=True)
            # create .csv with DROPS
            self.df_metrics_table.to_csv(os.path.join(self.outbound_dir, "df_metrics_DROP.csv"), index=True)
            # Set index of METRICS
            self.df_metrics_table.set_index(['QCDate', 'ProductLevel'], inplace=True)
            print(self.df_metrics_table)
            # Display Index
            print(self.df_metrics_table.index)
            #############################
            # create .csv with UNSORTED #
            #############################
            self.df_metrics_table.to_csv(os.path.join(self.outbound_dir, "df_metrics_UNSORTED.csv"), index=True)
            # SORT INDEX
            self.df_metrics_table.sort_index(inplace=True)
            ###########################
            # create .csv with SORTED #
            ###########################
            self.df_metrics_table.to_csv(os.path.join(self.outbound_dir, "df_metrics_SORTED.csv"), index=True)
            print("SORTED:\n")
            print(self.df_metrics_table.index)
            # SLICE & DICE THE DATAFRAME
            self.df_level_1s = self.df_metrics_table.loc[('2017-01-02', float(1))]
            print(self.df_level_1s)
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
        # }
    # }
    """
    
    """
    Referred to as "ProflowII" in SQL-Server
    """
    def pull_ProdflowII_table(self, table_name): # {
        # TRY THE FOLLOWING
        try: # {
            # CREATE CONNECTION STRING
            conn_str = str(
                r'DRIVER={ODBC Driver 17 for SQL Server};'
                r'SERVER=wtkngappflow1.is.agilent.net;'
                r'DATABASE=ProdFlowII_Prod;'
                r'Trusted_Connection=yes;'
            )
            # CREATE PYODBC CONNECTION
            cnxn_ProdflowII = pyodbc.connect(conn_str)
            # [2020-02-028]\\crsr_ProdflowII = cnxn_ProdflowII.cursor()
            # PERFORM SQL QUERY AND SET AS DATAFRAME
            df_ProdflowII_table = pd.read_sql_query(sql='SELECT * FROM ' + str(table_name),
                                                    con=cnxn_ProdflowII
                                                    )
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # { 
            print("Operation Completed Successfully...")
            return df_ProdflowII_table
        # } 
    # }
    
    """
    Referred as "Prodflow" in SQL-Server
    """
    def pull_ProdflowIII_table(self, table_name): # {
        # TRY THE FOLLOWING
        try: # {
            # CREATE CONNECTION STRING
            conn_str = str(
                r'DRIVER={ODBC Driver 17 for SQL Server};'
                r'SERVER=wtkngappflow1.is.agilent.net;'
                r'DATABASE=ProdFlow;'
                r'Trusted_Connection=yes;'
            )
            # CREATE PYODBC CONNECTION
            cnxn_ProdflowIII = pyodbc.connect(conn_str)
            # [2020-02-28]\\crsr_ProdflowIII = cnxn_ProdflowIII,cursor()
            # PERFORM SQL QUERY AND SET AS DATAFRAME
            df_ProdflowIII_table = pd.read_sql_query(sql='SELECT * FROM ' + str(table_name),
                                                     con=cnxn_ProdflowIII,
                                                     parse_dates=['QCDate']
                                                     )
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
            return df_ProdflowIII_table
        # }
    # }
    
    def create_metrics_table(self): # { 
        # TRY THE FOLLOWING
        try: # {
            # pull PRODUCTS TABLE
            self.df_products = self.pull_ProdflowII_table(table_name='Products')
            # RENAME ['Product#'] column to ["ProductNo']
            self.df_products.rename(columns={'Product#': 'ProductNo'}, inplace=True)
            print(self.df_products.info())
            # pull tblProdflow TABLE
            self.df_tblProdflow = self.pull_ProdflowIII_table(table_name='tblProdflow')
            print(self.df_tblProdflow.info())
            # CREATE METRICS TABLE
            df_metrics_table = pd.merge(self.df_products, self.df_tblProdflow, on='ProductNo', how='right')
        # }
        except: # { 
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
            return df_metrics_table
        # } 
    # }
# }

if __name__ == "__main__": # {
    # logger = Logger(logging_output_dir="").loggger
    window = tk.Tk()
    application = AgilentQCMetrics(root= window, the_logger=None)
    window.config()
    window.mainloop()
    # CALL METRICS CLASS WITH A RANGE OF 1 MONTH
    # [2020-02-28]\\test_metrics = AgilentQCMetrics(time_unit='M', time_value=1)
    # [2020-03-04]\\test_metrics = AgilentQCMetrics(day_range=30)
# }