# -*- coding: utf-8 -*-
"""
Created on Wed Mar 25 16:21:03 2020

QC_Metrics (turn around time)

TABLES USED:
    - tblProdflow
    - PRODUCTS

===================================================
- Takes the INPUT OF USER (month?)
- pulls tables from Prodflow
- Returns the average turn around time (cycle-days) 
  PER PRODUCT LEVEL, PER WEEK RANGE (yayyyy)
===================================================

04/07/20 - began implementing PDF export feature...

@author: derbates
"""

# IMPORT THE GOODS
import os, sys, time
from time import sleep
from pathlib import Path
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import seaborn as sns
import pandas as pd
import numpy as np
from pandas import Series, DataFrame
import logging
import pyodbc
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
from ttkthemes import ThemedStyle
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox, filedialog, commondialog

class Logger(): # {
    
    def __init__(self): # {
        pass
    # }
# }

class Agilent_QC_TR_Metrics(): # {
    
    user_name = str(os.getlogin()) # get username
    outbound_dir = "C:/data/outbound/" + str(pd.Timestamp.now())[:10]
    desktop_dir = "OneDrive - Agilent Technologies/Desktop"
    
    def __init__(self, root, the_logger): # {
        self.root = root
        self.root.title("QC Metrics - Cycle Time")
        self.root.geometry('315x200+300+300')
        self.root.resizable(width=False, height=False)
        # Get/Set USERNAME & DESKTOP DIRECTORIES
        self.user_name_dir = os.path.join("C:/Users/", self.user_name)
        self.desktop_dir = os.path.join(self.user_name_dir, self.desktop_dir)
        print("USER_DIR == " + str(self.user_name_dir))
        print("DESKTOP == " + str(self.desktop_dir))
        # CHECK IF DIRECTORIES ABOVE EXIST, IF NOT... CREATE THEM...
        self.desktop_folder = os.path.join(self.desktop_dir, str(pd.Timestamp.now())[:10])
        print(self.desktop_folder)
        if not os.path.exists(self.desktop_folder): # {
            # MAKE IT EXIST!
            os.makedirs(self.desktop_folder)
        # }
        # OVERWRITE DESKTOP VAR
        self.desktop_dir = self.desktop_folder
        # INITITALIZE UI
        self.create_gui(the_root = self.root)
    # }
    
    def create_gui(self, the_root): # {
        # TRY THE FOLLOWING
        try: # {
            self.create_ttk_styles(the_root=the_root)
            self.create_label_frame(the_root=the_root)
            self.create_main_frame(the_root=the_root)
        # }
        except: # {
            print("FAILED GUI")
        # }
    # }
    
    def create_ttk_styles(self, the_root): # {
        # TRY THE FOLLOWING
        try: # { 
            self.style = ThemedStyle(the_root)
            # STYLE THEME
            self.style.set_theme("kroc") # arc
        # }
        except: # {
            pass
        # }
    # }
    
    def create_label_frame(self, the_root): # {
        self.labelframe = ttk.Frame(the_root)
        self.labelframe.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    # }
    
    def create_main_frame(self, the_root): # {
        # TRY THE FOLLOWING
        try: # {
            self.mainframe = ttk.Frame(self.labelframe)
            self.mainframe.pack(anchor=tk.CENTER, fill=tk.BOTH, expand=True)
            # END DATE LABEL
            ttk.Label(master=self.mainframe, text="Enter Date: \n(YYYY-MM-DD)"
                      ).place(x=20, y=20) #.grid(row=0, col=0, padx=10, pady=10) 
            # CREATE END DATE TK VAR TO HOLD STR
            self.end_date = tk.StringVar(master=the_root, value=str(pd.Timestamp.now())[:10])
            # ENTRY FOR END DATE
            ttk.Entry(master=self.mainframe, textvariable=self.end_date
                      ).place(x=140, y=20) #.grid(row=0, col=1, padx=10, pady=10)
            # NUM OF DAYS LABEL
            ttk.Label(master=self.mainframe, text="# of days back: "
                      ).place(x=20, y=65)# .grid(row=1, col=0, padx=10, pady=10) 
            self.num_of_days = tk.IntVar(master=the_root, value=1)
            # NUM OF DAYS SPINBOX
            ttk.Spinbox(master=self.mainframe, to=365, from_=0, textvariable=self.num_of_days
                        ).place(x=140, y=55) #.grid(row=1, col=1, padx=10, pady=10)
            # CREATE STRING TO HOLD POSSIBLE FILENAME
            # [2020-03-08]\\self.period_str = tk.StringVar(master=the_root, value="Day")
            self.filename_str = tk.StringVar(master=the_root, value=str(self.end_date.get()))
            # FILENAME LABELS AND ENTRY BOX
            ttk.Label(master=self.mainframe, text="Folder Name:\tQCMetrics_"
                      ).place(x=10, y= 100)
            ttk.Entry(master=self.mainframe, textvariable=self.end_date, state=tk.DISABLED,
                      width=16
                      ).place(x=170, y= 100)
            # [2020-04-07]\\
            """
            ttk.Label(master=self.mainframe, text=".xlsx"
                      ).place(x=270, y= 100)
            """
            # BROWSE FOR EXPORT LOCATION BUTTON
            self.export_button = ttk.Button(master=self.mainframe, text="EXPORT TO DESKTOP",
                       command=self.determine_range, width=20).place(x=20, y = 140)
            # KEY-TYPE (radio-button)
            self.key_type_var = tk.StringVar(master=self.mainframe, value="ProductLevel")
            radio_type_1 = ttk.Radiobutton(master=self.mainframe,
                                           variable=self.key_type_var,
                                           value="ProductLevel", text="ProductLevel")
            radio_type_1.place(x=190, y = 140)
            radio_type_2 = ttk.Radiobutton(master=self.mainframe,
                                           variable=self.key_type_var,
                                           value="QCValidation", text="QCValidation")
            radio_type_2.place(x=190, y = 165)
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
            messagebox.showerror(title="ERROR!",
                                 message=typeE +
                                         "\n" + fileE +
                                         "\n" + lineE +
                                         "\n" + messageE)
        # }
    # }
    
    def determine_range(self): # {
        # TRY THE FOLLOWING
        try: # {
            print("DETERMINING...")
            print(self.end_date.get())
            print(self.num_of_days.get())
            print(self.key_type_var.get())
            # CALL RUN
            self.run(date_input=str(self.end_date.get()), day_range=self.num_of_days.get())
        # }
        except: # {
            pass
        # }
    # }
    
    def create_dashboard(self, cycle_time, n, date_range): # {
        # TRY THE FOLLOWING
        try: # {
            # CREATE DATAFRAME FROM IMPORT
            columns = ('Cycle Time(days)', 'n', 'Date Range')
            levels = pd.CategoricalIndex(['1', '2', '3'])
            self.df_dashboard = pd.DataFrame(data=None, index=levels, columns=columns)
            # CREATE COLUMN DATA
            self.df_dashboard['Cycle Time(days)'] = cycle_time
            self.df_dashboard['n'] = n
            self.df_dashboard['Date Range'] = date_range
            print(self.df_dashboard)
            """
            <<< SEABORN PLOT ?? >>>
            """
            seaborn_test = sns.lineplot(x='Date', y='Product Level(s)', data=self.df_dashboard,
                         kind='line', hue='Twin_Cities')
            print(seaborn_test)
            sea_figure = seaborn_test.get_figure()
            sea_figure.savefig(os.path.join(self.desktop_dir, "seaborn-test-"
                                            + str(pd.Timestamp.now())[:10]))
        # }
        except: # {
            print("FAILED CREATING DATAFRAME OF QC_DASHBOARD")
        # }
        else: # {
            print("Operation Completed Successfully...")
        # }
    # }
    
    def run(self, date_input, day_range): # {
        # TRY THE FOLLOWING
        try: # {
            # Create date (OFF INPUT)
            the_date = pd.Timestamp(ts_input=str(date_input))
            # create day range
            x_days_ago = the_date - timedelta(days = int(day_range))
            df_x_range = pd.date_range(start=x_days_ago, end=the_date, freq='D')
            print(list(df_x_range))
            # CREATE METRICS TABLE FROM CLASS METHOD
            # (returning only time-series we want??)
            self.df_metrics_table = self.create_metrics_table(
                time_start=x_days_ago, # 2019-11-01
                time_end=the_date) # 2020-03-01
            """
            << CHANGE DATA-TYPES OF COLUMNS >>
            """
            # change to 'DATETIME'
            self.df_metrics_table['AmpTimeIn'] = pd.to_datetime(self.df_metrics_table['AmpTimeIn'])
            self.df_metrics_table['AmpTimeOut'] = pd.to_datetime(self.df_metrics_table['AmpTimeOut'])
            # change to 'STRING'
            self.df_metrics_table['PfBatchID'] = self.df_metrics_table['PfBatchID'].astype('object')
            self.df_metrics_table['ProductID'] = self.df_metrics_table['ProductID'].astype('object')
            """
            << GRAB ONLY LEVELS 1, 2, and 3 >>
            """
            value_list= ['1.0', '2.0', '3.0']
            level_1 = ['1.0']
            level_2 = ['2.0']
            level_3 = ['3.0']
            # GRAB DATAFRAME ROWS where COLUMN contains certain values
            # >>> based on "self.key_type_var.get()"
            self.df_metrics_obv = self.df_metrics_table[self.df_metrics_table[str(self.key_type_var.get())].isin(value_list)]
            # EXPORT ??
            # [2020-03-30]\\self.df_metrics_obv.to_csv("c:/data/outbound/2020-03-30/df_metrics_obv.csv")
            # THEN GRAB DATAFRAME ROWS for EACH PRODUCT LEVEL
            self.df_level_1 = self.df_metrics_table[self.df_metrics_table[str(self.key_type_var.get())].isin(level_1)] # ProductLevel
            print(self.df_level_1.info())
            self.df_level_2 = self.df_metrics_table[self.df_metrics_table[str(self.key_type_var.get())].isin(level_2)] # ProductLevel
            print(self.df_level_2.info())
            self.df_level_3 = self.df_metrics_table[self.df_metrics_table[str(self.key_type_var.get())].isin(level_3)] # ProductLevel
            print(self.df_level_3.info())
            print("LEN OF (all levels) == " + str(len(self.df_metrics_obv)))
            print("LEN OF (#01 levels) == " + str(len(self.df_level_1)))
            print("LEN OF (#02 levels) == " + str(len(self.df_level_2)))
            print("LEN OF (#03 levels) == " + str(len(self.df_level_3)))
            # CREATE LIST TO HOLD NUMBERS (n)
            level_1_n = len(self.df_level_1)
            level_2_n = len(self.df_level_2)
            level_3_n = len(self.df_level_3)
            all_numbers = [level_1_n, level_2_n, level_3_n]
            # CREATE << GROUPBY >> DATAFRAME
            # [2020-03-30]\\self.df_groupby_levels = pd.DataFrame(data=self.df_total_levels.groupby(['QCDate', 'ProductLevel'])[['PfBatchID'].count()])
            self.ProdsPerDay = self.df_metrics_obv.groupby(['QCDate', str(self.key_type_var.get())])[['PfBatchID']].count()
            print("\nPRODS-PER-DAY INDEX:\n" + str(self.ProdsPerDay.index))
            """
            << FILL NA >>
            """
            self.ProdsPerDay.fillna(value=0, inplace=True)
            # RENAME COLUMNS
            # [2020-04-21]\\ count >> level
            self.ProdsPerDay.rename(columns={'PfBatchID':'Level'}, 
                                    inplace=True)
            #### CREATE ProdsPerDay THATS ***NOT*** stacked
            self.UnstackedProds = self.ProdsPerDay.unstack(level=1, fill_value=0)
            self.UnstackedProds.to_csv(os.path.join(self.desktop_dir, "UnstackedProds-"
                                                  + str(pd.Timestamp.now())[:10]
                                                  + ".csv"), index=True)
            """
            << PRODS PER DAY UNSTACK (same variable) >>
            """
            self.ProdsPerDay = self.ProdsPerDay.unstack(level=-1, fill_value=0)
            """
            <<< EXPORT TO DESKTOP >>>
            """
            filename_var = str("QC-ProdsPerDay-" + str(the_date)[:10] + ".csv")
            file_path = os.path.join(self.desktop_dir, str(filename_var))
            self.ProdsPerDay.to_csv(file_path)
            """
            <<< DETERMINE DAYS IN QC >>>
            """
            self.df_days_in_QC = pd.DataFrame(data=self.df_metrics_obv['PfBatchID'])
            # CREATE ANOTHA COLUMN
            self.df_days_in_QC['Date'] = self.df_metrics_obv['QCDate']
            # [2020-03-30]\\self.df_days_in_QC['ProductLevel'] = self.df_metrics_obv['ProductLevel']
            self.df_days_in_QC[str(self.key_type_var.get())] = self.df_metrics_obv[str(self.key_type_var.get())]
            self.df_days_in_QC['AmpDate'] = self.df_metrics_obv['AmpDate']
            ################
            # CALCULCATION #
            ################
            self.calculation = (self.df_days_in_QC['Date'] - self.df_days_in_QC['AmpDate'])
            # Assign Calculation COlumn
            self.df_days_in_QC['QC_days'] = self.calculation
            print(self.df_days_in_QC)
            #####################
            # CHANGE DATA-TYPES #
            #####################
            self.df_days_in_QC['QC_days'] = pd.to_timedelta(self.df_days_in_QC['QC_days'])
            ###########
            # DROP NA #
            ###########
            self.df_days_in_QC.dropna(subset=[str(self.key_type_var.get()), 'QC_days', 'AmpDate'],
                                      inplace=True)
            # CREATE COLUMN OF TYPE INT
            self.df_days_in_QC['Cycle_Time(days)'] = self.df_days_in_QC['QC_days'].dt.days
            print(self.df_days_in_QC)
            ################
            # CHANGE INDEX #
            ################
            self.df_days_in_QC.set_index(['Date'], drop=True, inplace=True)
            ##############
            # SORT INDEX #
            ##############
            self.df_days_in_QC.sort_index(inplace=True)
            """
            <<< ROLLING >>>
            """
            # Trail-rolling average transform [COLUMN]
            rolling = self.df_days_in_QC['Cycle_Time(days)'].rolling(window=1)
            rolling_mean = rolling.mean()
            # ASSIGN NEW COLUMN
            self.df_days_in_QC['rolling_avg'] = rolling_mean
            # [2020-03-31]\\
            """
            self.df_days_in_QC.to_csv(os.path.join(self.desktop_dir, "df_days_in_QC"
                                                   + str(pd.Timestamp.now())[:10]
                                                   + ".csv"))
            """
            """
            <<< GROUP-BY >>
            """
            # FORWARD_FILL
            # [2020-03-30]\\self.forward_fill = self.df_days_in_QC.groupby("ProductLevel")['Cycle_Time(days)'].resample("D").ffill()
            self.days_cycle = pd.DataFrame(data=self.df_days_in_QC.groupby(str(self.key_type_var.get()))['Cycle_Time(days)'].resample("D").mean()).reset_index()
            """
            <<< EXPORT TO DESKTOP >>>
            """
            filename_var = str("QC-CycleTime-" + str(the_date)[:10] + ".csv")
            file_path = os.path.join(self.desktop_dir, str(filename_var))
            self.days_cycle.to_csv(file_path, index=False)
            # EXPORT
            # [2020-03-31]\\
            """
            self.days_cycle.to_csv(os.path.join(self.desktop_dir, "QC-CycleTime-"
                                                + str(pd.Timestamp.now())[:10]
                                                + ".csv"))
            """
            # PLOT?
            names = self.df_days_in_QC[str(self.key_type_var.get())]
            values = self.days_cycle['Cycle_Time(days)']
            
            ########################
            # ['ProductLevel']
            ########################
            self.level_1s = self.days_cycle[self.days_cycle[str(self.key_type_var.get())].isin(level_1)]
            self.level_1s.dropna(axis=0, how='any', inplace=True)
            self.level_2s = self.days_cycle[self.days_cycle[str(self.key_type_var.get())].isin(level_2)]
            self.level_2s.dropna(axis=0, how='any', subset=['Cycle_Time(days)'], inplace=True)
            self.level_3s = self.days_cycle[self.days_cycle[str(self.key_type_var.get())].isin(level_3)]
            self.level_3s.dropna(axis=0, how='any', subset=['Cycle_Time(days)'], inplace=True)
            ########################
            # CALCULATE AVG AGAIN #
            ########################
            print("===========================\n\n")
            print("AVG OF level1 == " + str(self.level_1s['Cycle_Time(days)'].mean()))
            print("AVG OF level2 == " + str(self.level_2s['Cycle_Time(days)'].mean()))
            print("AVG of level3 == " + str(self.level_3s['Cycle_Time(days)'].mean()))
            level_1_avg = self.level_1s['Cycle_Time(days)'].mean()
            level_2_avg = self.level_2s['Cycle_Time(days)'].mean()
            level_3_avg = self.level_3s['Cycle_Time(days)'].mean()
            # CREATE LIST TO HOLD AVERAGES
            all_averages = [level_1_avg, level_2_avg, level_3_avg]
            print("ALL AVERAGES:\t" + str(all_averages))
            # PRINT ALL NUMBERS (generated above)
            print("ALL NUMBERS:\t" + str(all_numbers))
            # [2020-03-31]\\
            """
            # << EXPORT >>
            self.level_1s.to_csv(os.path.join(self.desktop_dir, "QC-CycleTime-Level1-"
                                              + str(pd.Timestamp.now())[:10]
                                              + ".csv"))
            self.level_2s.to_csv(os.path.join(self.desktop_dir, "QC-Cycle-Time-Level1-"
                                              + str(pd.Timestamp.now())[:10]
                                              + ".csv"))
            self.level_3s.to_csv(os.path.join(self.desktop_dir, "QC-CycleTime-Level3-"
                                              + str(pd.Timestamp.now())[:10]
                                              + ".csv"))
            """
            # GET STRINGS FOR DASHBOARD CREATION
            x_days_ago_str = str(x_days_ago)[:10]
            the_date_str = str(the_date)[:10]
            """
            [[[ CREATE DASHBOARD (csv) ]]]
            """
            self.create_dashboard(cycle_time=all_averages, n=all_numbers,
                                  date_range=str(x_days_ago_str + " to " + the_date_str))
            """
            [[[ EXPORT DASHBOARD (Csv) ]]]
            """
            filename_var = str("QC-Dashboard-" + str(the_date)[:10] + ".csv")
            file_path = os.path.join(self.desktop_dir, str(filename_var))
            self.df_dashboard.to_csv(file_path, index_label="Level")
            """
            [[[ CREATE LINE-CHART (dashboard) ]]]
            """
            self.create_dashboard_plot(the_dashboard_df=self.df_dashboard)
            """
            [ ============================================== ]
            """
            # create PLOTs off DataFrame
            dashboard_plot = self.df_dashboard.plot()
            # [2020-04-14]\\ cycletime_plot = self.days_cycle.plot()
            # cycletime_plot = self.df_days_in_QC.plot()
            prodspday_plot = self.ProdsPerDay.plot()
            # CREATE MATPLOT LIB FIGURE VARIABLES
            dashboard_fig = dashboard_plot.get_figure()
            #cycletime_fig = cycletime_plot.get_figure()
            prodspday_fig = prodspday_plot.get_figure()
            # SET TITLES?
            #dashboard_fig.title("QC Dashboard")
            #cycletime_fig.title("Average Turn-Around-Time (cycle time)")
            #prodspday_fig.title("Products Per Day")
            """
            plt.plot(self.df_dashboard['Date Range'], self.df_dashboard['n'], color='red', marker='o')
            plt.title('Total Products Tested Per Week Per Product Type', fontsize=10)
            plt.xlabel('Date Range', fontsize=10)
            plt.ylabel('Number of Products',fontsize=10)
            plt.grid(True)
            """
            # [2020-04-14]\\ts_str = str(pd.Timestamp.now())[:10]
            ts_str = str(the_date)[:10]
            ###################
            # CREATE TEMP DIR #
            ###################
            with tempfile.TemporaryDirectory() as directory_name: # {
                the_dir = Path(directory_name)
                print("TEMPORARY DIR == " + str(the_dir))
                dashboard_img_path = os.path.join(the_dir, "_dashboard_plot.png")
                #cycletime_img_path = os.path.join(the_dir, "_cycletime_plot.png")
                prodspday_img_path = os.path.join(the_dir, "_prdospday_plot.png")
                print("dashboard_plot_path == ")
                print("cycletime_plot_path == ")
                print("prodspday_plot_path == ")
                # SAVE PLOTS
                dashboard_fig.savefig(dashboard_img_path)
                # cycletime_fig.savefig(cycletime_img_path)
                prodspday_fig.savefig(prodspday_img_path)
                ###############################################
                # <<<<<<<<<< CALL WORKSHEET FUNCTION >>>>>> # 
                #################################################
                self.create_excel_workbook(the_date=ts_str, 
                                           graph_1_path=dashboard_img_path,
                                           #graph_2_path=cycletime_img_path,
                                           graph_2_path=prodspday_img_path)
            # }
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
        # }
    # }
    
    """
    TAKES IN: 
    (1) - date as string
    (2) - filename of OUTPUT file which will be saved at Desktop
    (3) - dataframe of *DATA* to put in sheet-1
    (4) - dataframe of *GRAPH* to put in sheet-2
    RETURNS: Nada, will have output an .xlsx file to location when complete
    """
    def create_excel_workbook(self, the_date, graph_1_path, graph_2_path): # {
        # TRY THE FOLLOWING
        try: # {
            print(self.desktop_dir)
            # CREATE STR VAR FOR filename convention
            ts_str = str(the_date)[:10]
            # CREATE FILENAME VAR
            filename_var = str("QC_TR_Metrics_" + str(ts_str) + ".xlsx")
            # CREATE FULL WORKBOOK PATH
            workbook_path = os.path.join(self.desktop_dir, str(filename_var))
            print("WORKBOOK PATH == " + str(workbook_path))
            # CREATE NEW WORKBOOK
            wb = Workbook()
            wb.save(workbook_path)
            # ADD SHEETS TO WORKBOOK 
            # DESIGNATE SHEET NAME AND POSITION
            sheet1 = wb.create_sheet('Graphs', 0)
            # ACTIVATE WORKSHEET
            active = wb['Graphs']
            # INSERT PLOT INTO WORKSHEET
            # Select Active sheet and cell reference
            img_graph_1 = Image(graph_1_path)
            active.add_image(img_graph_1, 'A1')
            # INSERT PLOT INTO WORKSHEET
            # Select Active sheet and cell reference
            img_graph_2 = Image(graph_2_path)
            active.add_image(img_graph_2, 'N1')
            """
            # INSERT PLOT INTO WORKSHEET
            # Select Active sheet and cell reference
            img_graph_3 = Image(graph_3_path)
            active.add_image(img_graph_3, 'Z1')
            """
            # SAVE WORKBOOK
            wb.save(workbook_path)
        # }
        except: # {
            pass
        # }
        else: # {
            print("Operation Completed Successfully...")
        # }
    # }
    
    """
    TAKES IN:
    (1) DataFrame of Dashboard
    """
    def create_dashboard_plot(self, the_dashboard_df): # {
        try: # {
            # CREATE EMPTY LISTS TO HOLD DATA
            idx_dates = []
            level_1 = []
            level_2 = []
            level_3 = []
            for row in self.ProdsPerDay.itertuples(index=True, name='Product'): # {
                print("INDEX == " + str(row[0]))
                print("LEVEL 1 COUNT == " + str(row[1]))
                print("LEVEL 2 COUNT == " + str(row[2]))
                print("LEVEL 3 COUNT == " + str(row[3]))
                idx_dates.append(str(row[0]))
                level_1.append(str(row[1]))
                level_2.append(str(row[2]))
                level_3.append(str(row[3]))
            # }
            # CREATE SERIES OFF OF LISTS
            s_level_1 = pd.Series(data=level_1, dtype=np.str)
            s_level_2 = pd.Series(data=level_2, dtype=np.str)
            s_level_3 = pd.Series(data=level_3, dtype=np.str)
            s_idx_dates = pd.Series(data=idx_dates, dtype=np.str)
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        ####################### 
        # PRODS PER DAY CHECK #
        #######################
        try: # {
            print(self.ProdsPerDay.head(10))
            single_column = self.ProdsPerDay.loc['2020-04-20']
            print("SINGLE COLUMN:\n" + str(single_column))
            print("TYPE == " + str(type(single_column)))
            """
            single_column.to_csv(os.path.join(self.desktop_dir, "column-"
                                                               + str(pd.Timestamp.now())[:10]
                                                               + ".csv"), index=False)
            """
            # TRY AND GET LEVELS FOR CERTAIN DAYS
            # LEVEL 1
            print("LEVEL 1:\t" + str(self.ProdsPerDay.loc['2020-04-20', '1']))
            # LEVEL 2
            print("LEVEL 2:\t" + str(self.ProdsPerDay.loc['2020-04-20', '2']))
            # LEVEL 3
            print("LEVEL 3:\t" + str(self.ProdsPerDay.loc['2020-04-20', '3']))
            print("\n ========================================== \n")
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        """
        ####################################################
        # TRY TO UNSTACK PREVIOUSLY MADE DATAFRAME
        try: # {
            # UNSTACK THE INDEX SO ITS JUST QCDATE ??? 
            df_unstacked = self.ProdsPerDay.unstack(level=-1)
            # LIST NUMBER OF COLUMN LEVELS
            print(len(df_unstacked.columns.levels))
            ##############
            # DROP LEVEL #
            ##############
            df_unstacked.columns = df_unstacked.columns.droplevel()
            # FILL NA
            # [2020-04-21]\\ df_unstacked.fillna(value=0, inplace=True)
            print("\n\n\t df_unstacked == " + str(df_unstacked.head()))
        # }
        except: # {
            print("\n <<< FAILED UNSTACKED >>> \n")
        # }
        """
        ##############################
        # TRY PULLING DATES AND SHIT #
        ##############################
        the_dashboard_df['Start_Date'] = the_dashboard_df['Date Range'].str[:10]
        the_dashboard_df['End_Date'] = the_dashboard_df['Date Range'].str[-10:]
        print(the_dashboard_df.head())
        try: # {
            # PULL SUb-STRING FROM DATA-COLUMN
            date_col_str = str(the_dashboard_df.iloc[0, 2])
            print("DATE-COL-STR == " + date_col_str)
            # pull START DATE from STR
            the_start_date = date_col_str[:10]
            # create index var to get end date out of STR
            end_date_idx = date_col_str.find('to ', 0)
            # pull END DATE from STR
            the_end_date = date_col_str[end_date_idx + 3:len(date_col_str)]
            print("START-DATE == " + str(the_start_date))
            print("END-DATE == " + str(the_end_date))
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        #################################
        # CREATE ACTUAL MATPLOTLIB PLOT #
        #################################
        try: # {
            """
            # CREATE DATE-RANGE/DATETIME-INDEX oFF column
            the_date_range = pd.date_range(start=the_start_date, 
                                           end=the_end_date,
                                           freq='D')
            print(the_date_range)
            """
            # CREATE EMPTY DATAFRAME (to be filled)
            the_result_df = pd.DataFrame(data=None, dtype=np.str)
            # ASSIGN SERIES CREATED ABOVE AS COLUMNS
            the_result_df['Level 1'] = s_level_1
            the_result_df['Level 2'] = s_level_2
            the_result_df['Level 3'] = s_level_3
            the_result_df['Date'] = s_idx_dates
            # SET INDEX TO DATE COLUMN
            the_result_df.set_index(['Date'], drop=True, inplace=True)
            # SORT INDEX
            the_result_df.sort_index(inplace=True)
            print(the_result_df)
            the_result_df.to_csv(os.path.join(self.desktop_dir, "test-result-"
                                              + str(pd.Timestamp.now())[:10]
                                              + ".csv"), index=True)
            test_plot = the_result_df.plot()
            test_figure = test_plot.get_figure()
            test_figure.savefig(os.path.join(self.desktop_dir, "test-figure-"
                                             + str(pd.Timestamp.now())[:10]))
            """
            # CREATE "empty" DATAFRAME, fill with all values just made
            the_result_df = pd.DataFrame(data=None, index=the_date_range)
            print(the_result_df)
            """
            time_1 = the_result_df.index
            p_level_1 = the_result_df['Level 1']
            p_level_2 = the_result_df['Level 2']
            p_level_3 = the_result_df['Level 3']
            #level_1_plot
            plt.plot(time_1, p_level_1, color='blue', marker='o')
            #level_2_plot
            plt.plot(time_1, p_level_2, color='orange', marker='o')
            #level_3_plot
            plt.plot(time_1, p_level_3, color='gray', marker='o')
            plt.title('Products Per Week Per Product Type', fontsize=14)
            plt.xlabel('Time', fontsize=14)
            plt.ylabel('Product Levels', fontsize=14)
            plt.grid(True)
            plt.show()
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
        # }
    # }
    
    def calculate_moving_average(self, dataframe_to_avg, columns_to_avg): # {
        # TRY THE FOLLOWING
        try: # {
            weighted_avg = (
                dataframe_to_avg[str(columns_to_avg[0])] 
                * dataframe_to_avg[str(columns_to_avg[1])].sum() 
                / dataframe_to_avg(str(columns_to_avg[2])).sum()
                )
            print("\n\tWEIGHTED AVERAGE:\n\t" + str(weighted_avg))
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
            return weighted_avg
        # }
    # }
    
    """
    Referred to as "ProdflowII" in SQL-Server
    """
    def pull_ProdflowII_table(self, table_name): # {
        # TRY THE FOLLOWING
        try: # {
            # CREATE CONNECTION STRING
            conn_str = str(
                r'DRIVER={ODBC Driver 17 for SQL Server};'
                r'SERVER=wtkngappflow1.is.agilent.net;'
                r'DATABASE=ProdFlowII_Prod;'
                r'Trusted_Connection=yes;'
            )
            # CREATE PYODBC CONNECTION
            cnxn_ProdflowII = pyodbc.connect(conn_str)
            # [2020-02-028]\\crsr_ProdflowII = cnxn_ProdflowII.cursor()
            # PERFORM SQL QUERY AND SET AS DATAFRAME
            df_ProdflowII_table = pd.read_sql_query(sql='SELECT * FROM ' + str(table_name),
                                                    con=cnxn_ProdflowII
                                                    )
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
            return df_ProdflowII_table
        # }
    # }
    
    """
    Referred to as "Prodflow" in SQL-Server
    """
    def pull_ProdflowIII_table(self, table_name): # {
        # TRY THE FOLLOWING
        try: # {
            # CREATE CONNECTION STRING
            conn_str = str(
                r'DRIVER={ODBC Driver 17 for SQL Server};'
                r'SERVER=wtkngappflow1.is.agilent.net;'
                r'DATABASE=ProdFlow;'
                r'Trusted_Connection=yes;'
            )
            # CREATE PYODBC CONNECTION
            cnxn_ProdflowIII = pyodbc.connect(conn_str)
            # [2020-02-28]\\crsr_ProdflowIII = cnxn_ProdflowIII,cursor()
            # PERFORM SQL QUERY AND SET AS DATAFRAME
            df_ProdflowIII_table = pd.read_sql_query(sql='SELECT * FROM ' + str(table_name),
                                                     con=cnxn_ProdflowIII,
                                                     parse_dates=['QCDate',
                                                                  'AmpDate',
                                                                  'OriginationDate',
                                                                  'EntryDate'
                                                                  ]
                                                     )
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
           print("Operation Completed Successfully...")
           return df_ProdflowIII_table
        # } 
    # }
    
    def create_metrics_table(self, time_start, time_end): # {
        # TRY THE FOLLOWING
        try: # {
            # pull PRODUCTS TABLE
            self.df_products = self.pull_ProdflowII_table(table_name='Products')
            # RENAME ["Product#'] column to ["ProductNo"]
            self.df_products.rename(columns={'Product#':'ProductNo'}, 
                                    inplace=True)
            print(self.df_products.info())
            # pull ORDERS TABLE
            self.df_orders = self.pull_ProdflowII_table(table_name='Orders')
            print(self.df_orders.info())
            # pull tblProdflow TABLE
            self.df_tblProdflow = self.pull_ProdflowIII_table(table_name='tblProdflow')
            print(self.df_tblProdflow.info())
            # CREATE METRICS TABLE
            df_metrics_table = pd.merge(self.df_products, self.df_tblProdflow, 
                                        on='ProductNo', how='right')
            # DROP ALL ROWS WITHOUT A 'QCDate' & 'ProductLevel'
            df_metrics_table.dropna(axis=0, subset=['QCDate', str(self.key_type_var.get())], 
                                    how='any', inplace=True)
            # ONLY USE ROWS THAT ARE WITHIN SPECIFIED TIME FRAME
            # time_start, time_end
            df_QC_metrics = df_metrics_table[(df_metrics_table['QCDate'] > time_start) & (df_metrics_table['QCDate'] < time_end)] 
        # }
        except: # {
            errorMessage = str(sys.exc_info()[0]) + "\n"
            errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
            errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            typeE = str("TYPE : " + str(exc_type))
            fileE = str("FILE : " + str(fname))
            lineE = str("LINE : " + str(exc_tb.tb_lineno))
            messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
            print("\n" + typeE + 
                  "\n" + fileE + 
                  "\n" + lineE + 
                  "\n" + messageE)
        # }
        else: # {
            print("Operation Completed Successfully...")
            # [2020-03-27]\\return df_metrics_table
            return df_QC_metrics
        # }
    # }
    
# }

def main(): # { 
    # TRY THE FOLLOWING
    try: # {
        window = tk.Tk()
        application = Agilent_QC_TR_Metrics(root = window, the_logger=None)
        window.config()
        window.mainloop()
    # }
    except: # {
        errorMessage = str(sys.exc_info()[0]) + "\n"
        errorMessage = errorMessage + str(sys.exc_info()[1]) + "\n\t\t"
        errorMessage = errorMessage + str(sys.exc_info()[2]) + "\n"
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        typeE = str("TYPE : " + str(exc_type))
        fileE = str("FILE : " + str(fname))
        lineE = str("LINE : " + str(exc_tb.tb_lineno))
        messageE = str("MESG : " + "\n\n" + str(errorMessage) + "\n")
        print("\n" + typeE + 
              "\n" + fileE + 
              "\n" + lineE + 
              "\n" + messageE)
    # }
    else: # {
        print("Operation Completed Successfully...")
    # }
# }


if __name__ == "__main__": # {
    # call main function
    main()
# }